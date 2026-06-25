import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil

# 文件路径
source_file = r'E:\B01Python\PythonProject\A05_fx_rate\fx_rate\fx_rate.csv'
target_file = r'E:\2026年QDII LOF基金估值偏差记录.xlsm'


def update_usd_cny_safe():
    """
    安全更新USD/CNY汇率，保留所有格式和宏
    """
    try:
        # 1. 先备份原文件
        backup_file = target_file.replace('.xlsm', '_backup.xlsm')
        print(f"正在备份原文件到: {backup_file}")
        shutil.copy2(target_file, backup_file)

        # 2. 读取源CSV文件
        print("正在读取源文件...")
        df_source = pd.read_csv(source_file)
        df_source['Date'] = pd.to_datetime(df_source['Date'])

        # 构建汇率字典
        rate_dict = {}
        for _, row in df_source.iterrows():
            date_key = row['Date'].date()
            rate_dict[date_key] = row['USD/CNY']
            # 也存储字符串格式
            rate_dict[row['Date'].strftime('%Y-%m-%d')] = row['USD/CNY']
            rate_dict[row['Date'].strftime('%Y/%m/%d')] = row['USD/CNY']

        print(f"已加载 {len(df_source)} 条汇率数据")
        print(f"汇率日期范围: {df_source['Date'].min().date()} 至 {df_source['Date'].max().date()}")

        # 3. 使用openpyxl直接操作，保留所有格式
        print("正在打开目标文件...")
        wb = load_workbook(target_file, keep_vba=True, data_only=False)

        total_updated = 0
        processed_sheets = []

        # 4. 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n处理工作表: {sheet_name}")

            # 查找日期列和USD/CNY列
            date_col_idx = None
            rate_col_idx = None

            # 扫描第一行找列
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx)
                if cell.value:
                    col_name = str(cell.value).strip()
                    if col_name.upper() in ['DATE', '日期']:
                        date_col_idx = col_idx
                        print(f"  找到日期列: 第{col_idx}列 '{col_name}'")
                    if 'USD/CNY' in col_name.upper():
                        rate_col_idx = col_idx
                        print(f"  找到USD/CNY列: 第{col_idx}列 '{col_name}'")

            if date_col_idx is None:
                print(f"  ⚠️ 跳过: 未找到日期列")
                continue

            if rate_col_idx is None:
                print(f"  ⚠️ 跳过: 未找到USD/CNY列")
                continue

            # 处理数据行
            updated_count = 0
            total_rows = 0

            for row_idx in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row_idx, column=date_col_idx)

                # 如果日期单元格为空，跳过
                if date_cell.value is None:
                    continue

                total_rows += 1

                try:
                    # 获取日期值
                    date_val = date_cell.value
                    date_key = None

                    # 处理不同类型的日期值
                    if isinstance(date_val, datetime):
                        date_key = date_val.date()
                    elif isinstance(date_val, str):
                        # 如果是公式（以=开头），先计算
                        if date_val.startswith('='):
                            # openpyxl无法计算公式，跳过
                            continue
                        # 尝试解析日期字符串
                        try:
                            # 尝试多种格式
                            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%m/%d/%Y', '%d/%m/%Y']:
                                try:
                                    dt = datetime.strptime(date_val, fmt)
                                    date_key = dt.date()
                                    break
                                except:
                                    continue
                            if date_key is None:
                                dt = pd.to_datetime(date_val)
                                date_key = dt.date()
                        except:
                            continue
                    else:
                        # 其他类型尝试转换
                        try:
                            dt = pd.to_datetime(date_val)
                            date_key = dt.date()
                        except:
                            continue

                    # 查找匹配的汇率
                    if date_key:
                        # 先尝试用date对象匹配
                        if date_key in rate_dict:
                            new_rate = rate_dict[date_key]
                            ws.cell(row=row_idx, column=rate_col_idx, value=new_rate)
                            updated_count += 1
                            if updated_count <= 3:
                                print(f"  行{row_idx}: {date_key} -> {new_rate}")
                        else:
                            # 尝试用字符串匹配
                            date_str1 = date_key.strftime('%Y-%m-%d')
                            date_str2 = date_key.strftime('%Y/%m/%d')
                            if date_str1 in rate_dict:
                                ws.cell(row=row_idx, column=rate_col_idx, value=rate_dict[date_str1])
                                updated_count += 1
                            elif date_str2 in rate_dict:
                                ws.cell(row=row_idx, column=rate_col_idx, value=rate_dict[date_str2])
                                updated_count += 1

                except Exception as e:
                    # 单个日期处理失败，跳过
                    continue

            print(f"  总计 {total_rows} 行有日期，成功更新 {updated_count} 行")
            total_updated += updated_count
            processed_sheets.append(sheet_name)

        # 5. 保存文件
        print("\n正在保存文件...")
        wb.save(target_file)
        wb.close()

        print("\n" + "=" * 50)
        print(f"✅ 更新完成！")
        print(f"  处理工作表: {processed_sheets}")
        print(f"  总共更新: {total_updated} 行")
        print(f"  备份文件: {backup_file}")
        print("=" * 50)

        return True

    except PermissionError:
        print("❌ 权限错误！请确保目标文件未被Excel打开")
        print("   请关闭Excel文件后重新运行")
        return False
    except Exception as e:
        print(f"❌ 发生错误: {e}")
        import traceback
        traceback.print_exc()

        # 如果有备份，提示恢复
        if os.path.exists(backup_file):
            print(f"\n⚠️ 如果文件损坏，可以从备份恢复: {backup_file}")
        return False


if __name__ == "__main__":
    if not os.path.exists(source_file):
        print(f"❌ 源文件不存在: {source_file}")
    elif not os.path.exists(target_file):
        print(f"❌ 目标文件不存在: {target_file}")
    else:
        update_usd_cny_safe()