import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil
import warnings

warnings.filterwarnings('ignore')


class FundDataUpdater:
    def __init__(self, target_file, backup_file):
        self.target_file = target_file
        self.backup_file = backup_file
        self.source_dirs = {
            'fund_nav': r'E:\B01Python\PythonProject\A01_fund_nav_data\fund_nav_data',
            'index_data': r'E:\B01Python\PythonProject\A02_index_and_fund_close_data\index_data',
            'fund_share': r'E:\B01Python\PythonProject\A03_fund_share_data\fund_share_data',
            'fx_rate': r'E:\B01Python\PythonProject\A05_fx_rate\fx_rate\fx_rate.csv',
            'xop_nav': r'E:\B01Python\PythonProject\A06_xop\xop_daily_nav.csv',
            'xop_price': r'E:\B01Python\PythonProject\A06_xop\xop_gld_history.csv'
        }

    def backup_target_file(self):
        """备份目标文件"""
        try:
            shutil.copy2(self.target_file, self.backup_file)
            print(f"备份文件已创建: {self.backup_file}")
            return True
        except Exception as e:
            print(f"备份文件失败: {e}")
            return False

    def find_date_column(self, df):
        """查找日期列"""
        date_keywords = ['date', '日期', 'Date', 'DATE', '时间', 'Time']
        for col in df.columns:
            if col.strip() in date_keywords or col.strip().lower() in [k.lower() for k in date_keywords]:
                return col
        # 如果没有找到，尝试查找包含date的列名
        for col in df.columns:
            if 'date' in col.lower() or '时间' in col:
                return col
        return None

    def load_source_data(self):
        """加载所有源数据"""
        source_data = {}

        # 1. 加载基金净值数据
        fund_nav_data = {}
        nav_dir = self.source_dirs['fund_nav']
        if os.path.exists(nav_dir):
            for file in os.listdir(nav_dir):
                if file.endswith('.csv'):
                    try:
                        fund_code = file.replace('.csv', '')
                        df = pd.read_csv(os.path.join(nav_dir, file))

                        # 查找日期列
                        date_col = self.find_date_column(df)
                        if date_col:
                            df[date_col] = pd.to_datetime(df[date_col])
                            # 查找nav列
                            nav_col = None
                            for col in df.columns:
                                if col.strip().lower() in ['nav', '净值']:
                                    nav_col = col
                                    break
                            if nav_col:
                                fund_nav_data[fund_code] = df.set_index(date_col)[nav_col]
                            else:
                                print(f"  警告: {file} 中未找到nav列")
                        else:
                            print(f"  警告: {file} 中未找到日期列")
                    except Exception as e:
                        print(f"  警告: 读取 {file} 时出错: {e}")
        source_data['fund_nav'] = fund_nav_data
        print(f"加载基金净值数据: {len(fund_nav_data)} 个基金")

        # 2. 加载指数和基金收盘数据
        index_data = {}
        index_dir = self.source_dirs['index_data']
        if os.path.exists(index_dir):
            for file in os.listdir(index_dir):
                if file.endswith('.csv'):
                    try:
                        fund_code = file[:6]  # 取前6位作为基金代码
                        df = pd.read_csv(os.path.join(index_dir, file))

                        # 查找日期列
                        date_col = self.find_date_column(df)
                        if date_col:
                            df[date_col] = pd.to_datetime(df[date_col])
                            index_data[fund_code] = df.set_index(date_col)
                        else:
                            print(f"  警告: {file} 中未找到日期列")
                    except Exception as e:
                        print(f"  警告: 读取 {file} 时出错: {e}")
        source_data['index_data'] = index_data
        print(f"加载指数和基金收盘数据: {len(index_data)} 个基金")

        # 3. 加载基金份额数据
        fund_share_data = {}
        share_dir = self.source_dirs['fund_share']
        if os.path.exists(share_dir):
            for file in os.listdir(share_dir):
                if file.endswith('.csv'):
                    try:
                        fund_code = file.replace('.csv', '')
                        df = pd.read_csv(os.path.join(share_dir, file))

                        # 查找日期列
                        date_col = self.find_date_column(df)
                        if date_col:
                            df[date_col] = pd.to_datetime(df[date_col])
                            # 查找share列
                            share_col = None
                            for col in df.columns:
                                if col.strip().lower() in ['share', '份额']:
                                    share_col = col
                                    break
                            if share_col:
                                fund_share_data[fund_code] = df.set_index(date_col)[share_col]
                            else:
                                print(f"  警告: {file} 中未找到share列")
                        else:
                            print(f"  警告: {file} 中未找到日期列")
                    except Exception as e:
                        print(f"  警告: 读取 {file} 时出错: {e}")
        source_data['fund_share'] = fund_share_data
        print(f"加载基金份额数据: {len(fund_share_data)} 个基金")

        # 4. 加载汇率数据
        fx_file = self.source_dirs['fx_rate']
        if os.path.exists(fx_file):
            try:
                df = pd.read_csv(fx_file)
                date_col = self.find_date_column(df)
                if date_col:
                    df[date_col] = pd.to_datetime(df[date_col])
                    source_data['fx_rate'] = df.set_index(date_col)
                    print(f"加载汇率数据: {len(df)} 条记录")
                else:
                    print(f"警告: 汇率文件中未找到日期列")
            except Exception as e:
                print(f"警告: 读取汇率文件时出错: {e}")

        # 5. 加载XOP净值数据
        xop_nav_file = self.source_dirs['xop_nav']
        if os.path.exists(xop_nav_file):
            try:
                df = pd.read_csv(xop_nav_file)
                date_col = self.find_date_column(df)
                if date_col:
                    df[date_col] = pd.to_datetime(df[date_col])
                    nav_col = None
                    for col in df.columns:
                        if col.strip().lower() in ['nav', '净值']:
                            nav_col = col
                            break
                    if nav_col:
                        source_data['xop_nav'] = df.set_index(date_col)[nav_col]
                        print(f"加载XOP净值数据: {len(df)} 条记录")
                    else:
                        print(f"警告: XOP净值文件中未找到nav列")
                else:
                    print(f"警告: XOP净值文件中未找到日期列")
            except Exception as e:
                print(f"警告: 读取XOP净值文件时出错: {e}")

        # 6. 加载XOP价格数据
        xop_price_file = self.source_dirs['xop_price']
        if os.path.exists(xop_price_file):
            try:
                df = pd.read_csv(xop_price_file)
                date_col = self.find_date_column(df)
                if date_col:
                    df[date_col] = pd.to_datetime(df[date_col])
                    source_data['xop_price'] = df.set_index(date_col)
                    print(f"加载XOP价格数据: {len(df)} 条记录")
                    print(f"  XOP价格数据列: {list(df.columns)}")
                else:
                    print(f"警告: XOP价格文件中未找到日期列")
            except Exception as e:
                print(f"警告: 读取XOP价格文件时出错: {e}")

        return source_data

    def safe_float_convert(self, value):
        """安全转换为浮点数"""
        try:
            if pd.isna(value) or value is None:
                return None
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                value = value.replace(',', '').strip()
                if value == '':
                    return None
                return float(value)
            return float(value)
        except:
            return None

    def update_excel_file(self, source_data):
        """更新Excel文件"""
        try:
            # 加载工作簿
            wb = load_workbook(self.target_file, keep_vba=True)

            # 获取所有工作表
            sheet_names = wb.sheetnames

            # 处理每个工作表
            for sheet_name in sheet_names:
                print(f"\n处理工作表: {sheet_name}")
                ws = wb[sheet_name]

                # 获取工作表表头
                headers = {}
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers[str(cell_value).upper().strip()] = col

                print(f"找到表头: {list(headers.keys())[:10]}...")

                # 获取日期列
                date_col = None
                for header, col in headers.items():
                    if header in ['DATE', '日期', 'DATE(日期)']:
                        date_col = col
                        break

                if not date_col:
                    print(f"  工作表 {sheet_name} 没有日期列，跳过")
                    continue

                # 获取所有日期和数据（从第2行开始，即数据行第1行）
                dates = []
                for row in range(2, ws.max_row + 1):
                    date_cell = ws.cell(row=row, column=date_col)
                    if date_cell.value:
                        try:
                            date_val = pd.to_datetime(date_cell.value)
                            dates.append((row, date_val))
                        except:
                            continue

                if not dates:
                    print(f"  工作表 {sheet_name} 没有有效日期，跳过")
                    continue

                print(f"  找到 {len(dates)} 行数据")

                # 1. 更新NAV数据
                if 'NAV' in headers:
                    nav_col = headers['NAV']
                    fund_code = sheet_name
                    if fund_code in source_data['fund_nav']:
                        nav_data = source_data['fund_nav'][fund_code]
                        updated = 0
                        for row, date_val in dates:
                            if date_val in nav_data.index:
                                value = self.safe_float_convert(nav_data[date_val])
                                if value is not None:
                                    ws.cell(row=row, column=nav_col, value=value)
                                    updated += 1
                        print(f"  更新NAV数据: {updated} 条")
                    else:
                        print(f"  未找到基金 {fund_code} 的净值数据")

                # 2. 更新CLOSE和Amount数据
                if sheet_name in source_data['index_data']:
                    index_df = source_data['index_data'][sheet_name]

                    # 更新CLOSE
                    if 'CLOSE' in headers:
                        close_col = headers['CLOSE']
                        # 查找fund_close列
                        fund_close_col = None
                        for col in index_df.columns:
                            if col.strip().lower() in ['fund_close', 'fund close', '基金收盘']:
                                fund_close_col = col
                                break
                        if fund_close_col:
                            updated = 0
                            for row, date_val in dates:
                                if date_val in index_df.index:
                                    value = self.safe_float_convert(index_df.loc[date_val, fund_close_col])
                                    if value is not None:
                                        ws.cell(row=row, column=close_col, value=value)
                                        updated += 1
                            print(f"  更新CLOSE数据: {updated} 条")

                    # 更新Amount
                    if 'AMOUNT' in headers:
                        amount_col = headers['AMOUNT']
                        # 查找fund_amount列
                        fund_amount_col = None
                        for col in index_df.columns:
                            if col.strip().lower() in ['fund_amount', 'fund amount', '基金金额']:
                                fund_amount_col = col
                                break
                        if fund_amount_col:
                            updated = 0
                            for row, date_val in dates:
                                if date_val in index_df.index:
                                    value = self.safe_float_convert(index_df.loc[date_val, fund_amount_col])
                                    if value is not None:
                                        ws.cell(row=row, column=amount_col, value=value)
                                        updated += 1
                            print(f"  更新Amount数据: {updated} 条")

                # 3. 更新share数据
                if 'SHARE' in headers:
                    share_col = headers['SHARE']
                    fund_code = sheet_name
                    if fund_code in source_data['fund_share']:
                        share_data = source_data['fund_share'][fund_code]
                        updated = 0
                        for row, date_val in dates:
                            if date_val in share_data.index:
                                value = self.safe_float_convert(share_data[date_val])
                                if value is not None:
                                    ws.cell(row=row, column=share_col, value=value)
                                    updated += 1
                        print(f"  更新share数据: {updated} 条")
                    else:
                        print(f"  未找到基金 {fund_code} 的份额数据")

                # 4. 更新USD/CNY汇率数据
                if 'USD/CNY' in headers or 'USD_CNY' in headers:
                    fx_col = None
                    for header, col in headers.items():
                        if header in ['USD/CNY', 'USD_CNY']:
                            fx_col = col
                            break

                    if fx_col and 'fx_rate' in source_data:
                        fx_data = source_data['fx_rate']
                        # 查找USD/CNY列
                        usd_cny_col = None
                        for col in fx_data.columns:
                            if col.strip().lower() in ['usd/cny', 'usd_cny', '美元/人民币']:
                                usd_cny_col = col
                                break
                        if usd_cny_col:
                            updated = 0
                            for row, date_val in dates:
                                if date_val in fx_data.index:
                                    value = self.safe_float_convert(fx_data.loc[date_val, usd_cny_col])
                                    if value is not None:
                                        ws.cell(row=row, column=fx_col, value=value)
                                        updated += 1
                            print(f"  更新USD/CNY数据: {updated} 条")

                # 5. 更新XOP_NAV数据 (仅162411工作表)
                if sheet_name == '162411' and 'XOP_NAV' in headers:
                    xop_nav_col = headers['XOP_NAV']
                    if 'xop_nav' in source_data:
                        xop_nav_data = source_data['xop_nav']
                        updated = 0
                        for row, date_val in dates:
                            if date_val in xop_nav_data.index:
                                value = self.safe_float_convert(xop_nav_data[date_val])
                                if value is not None:
                                    ws.cell(row=row, column=xop_nav_col, value=value)
                                    updated += 1
                        print(f"  更新XOP_NAV数据: {updated} 条")

                # 6. 特殊处理XOP_price列 (162411工作表)
                if sheet_name == '162411' and 'XOP_PRICE' in headers:
                    xop_price_col = headers['XOP_PRICE']
                    if 'xop_price' in source_data:
                        xop_df = source_data['xop_price']
                        # 查找XOP列
                        xop_col = None
                        for col in xop_df.columns:
                            if col.strip().upper() == 'XOP':
                                xop_col = col
                                break
                        if xop_col:
                            updated = 0
                            for row, date_val in dates:
                                if date_val in xop_df.index:
                                    value = self.safe_float_convert(xop_df.loc[date_val, xop_col])
                                    if value is not None:
                                        ws.cell(row=row, column=xop_price_col, value=value)
                                        updated += 1
                            print(f"  更新XOP_PRICE数据: {updated} 条")
                        else:
                            print(f"  警告: XOP价格数据中未找到XOP列")
                    else:
                        print(f"  警告: 未加载XOP价格数据")

                # 7. 更新其他匹配的列（所有工作表）
                if 'xop_price' in source_data:
                    xop_df = source_data['xop_price']

                    # 检查所有可能的列匹配（排除已处理的列）
                    for header, col in headers.items():
                        # 跳过已经处理过的列
                        if header in ['DATE', 'NAV', 'CLOSE', 'AMOUNT', 'SHARE', 'USD/CNY', 'USD_CNY',
                                      'XOP_NAV', 'XOP_PRICE', 'ADD SHARE', 'P', 'EST', 'IOPV', 'PM', 'IPM', 'BR',
                                      'BACKTEST']:
                            continue

                        # 在XOP数据中查找匹配的列
                        for xop_col in xop_df.columns:
                            if header.upper() == xop_col.upper():
                                updated = 0
                                for row, date_val in dates:
                                    if date_val in xop_df.index:
                                        value = self.safe_float_convert(xop_df.loc[date_val, xop_col])
                                        if value is not None:
                                            ws.cell(row=row, column=col, value=value)
                                            updated += 1
                                if updated > 0:
                                    print(f"  更新{header}数据: {updated} 条")
                                break

                # 8. 计算add share列（从第3行开始，即数据行第2行）
                if 'ADD SHARE' in headers and 'SHARE' in headers:
                    add_share_col = headers['ADD SHARE']
                    share_col = headers['SHARE']

                    # 收集所有日期和share数据（从第2行开始）
                    share_data = []
                    for row, date_val in dates:
                        share_val = ws.cell(row=row, column=share_col).value
                        if share_val is not None and pd.notna(share_val):
                            try:
                                share_data.append((row, float(share_val)))
                            except:
                                continue

                    # 从第3行开始计算add share（数据行的第2行）
                    # 表头是第1行，数据从第2行开始，所以第3行是数据行的第2行
                    if len(share_data) >= 3:  # 至少有3行数据才计算
                        # 从索引1开始（即数据行的第2行，对应Excel的第3行）
                        for i in range(1, len(share_data)):
                            row, current_share = share_data[i]
                            prev_row, prev_share = share_data[i - 1][0], share_data[i - 1][1]
                            # 只有当前行不是第2行（Excel行号>2）时才计算
                            if row > 2:  # Excel行号从1开始，第2行是数据第1行
                                add_share_val = (current_share - prev_share) / 10000
                                ws.cell(row=row, column=add_share_col, value=add_share_val)
                        print(f"  计算ADD SHARE数据: {len(share_data) - 1} 条")

            # 保存工作簿
            wb.save(self.target_file)
            print(f"\n数据更新完成，文件已保存: {self.target_file}")

        except Exception as e:
            print(f"更新Excel文件时出错: {e}")
            import traceback
            traceback.print_exc()
            raise

    def run(self):
        """运行整个更新过程"""
        print("=" * 60)
        print("开始更新基金数据...")
        print("=" * 60)

        # 1. 备份目标文件
        if not self.backup_target_file():
            print("备份失败，程序终止")
            return False

        # 2. 加载所有源数据
        source_data = self.load_source_data()

        # 3. 更新Excel文件
        try:
            self.update_excel_file(source_data)
            print("\n数据更新成功完成！")
            return True
        except Exception as e:
            print(f"\n数据更新失败: {e}")
            import traceback
            traceback.print_exc()
            return False


def main():
    target_file = r'E:\B01Python\PythonProject\to_github\D03_write_data_to_xlsx\QDII LOF基金估值偏差记录.xlsm'
    backup_file = r'E:\B01Python\PythonProject\to_github\D03_write_data_to_xlsx\QDII LOF基金估值偏差记录_backup.xlsm'

    updater = FundDataUpdater(target_file, backup_file)
    success = updater.run()

    if success:
        print("\n程序执行完成！")
    else:
        print("\n程序执行失败，请检查错误信息。")


if __name__ == "__main__":
    main()